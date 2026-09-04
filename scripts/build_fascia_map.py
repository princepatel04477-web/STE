"""
Builds public/assets/fascia_map.json: unitId -> up to 2 fascia (signboard)
names, sourced only from live Supabase (exhibitors.fascia_names_json, what
each exhibitor typed into the portal themselves) and
STE_2026_Admin_Master_Report_For_facia-names.xlsx (the organisers' own
reconciled sheet). Stall numbers come from master_exhibitor_stalls.csv - the
consolidated, alias-resolved live draw, not any static planning file.

Priority when both sources have a name for the same firm: Supabase, because
it is what the exhibitor is looking at on their own dashboard right now. The
admin sheet fills the gap when Supabase has nothing, and brand_name is the
last resort so nothing renders blank.

A bay with no live draw at all is only marked (Vacant) if truly nobody is
down for it. Otherwise src/data/stallAllotment2026.ts - the organisers'
static plan - names whoever the bay is earmarked for, so a firm that simply
has not opened the Lucky Box yet still shows a name instead of "(Vacant)".
This fallback is skipped for a unit whose planned brand has already drawn a
*different* real stall live (e.g. Nidhivan/Yogayaa was planned for 62 but
actually drew 53) - showing them again on 62 would just be wrong, and 62 has
nobody on it any more.
"""
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(r"C:\Users\rebel\OneDrive\Documents\Projects\STE-feature-clean-code-assets")
# A session-scoped temp path here would rot the moment that session ends -
# this one already had, pointing at a scratchpad no session could still
# reach. /scratch is the repo's own gitignored working directory instead, so
# a fresh exhibitors_fascia.json dropped there survives across sessions.
SCRATCH = ROOT / "scratch"


def clean_names(raw):
    seen = set()
    out = []
    for n in raw:
        s = str(n or "").strip()
        if not s:
            continue
        norm = s.lower()
        if norm in seen:
            continue
        seen.add(norm)
        out.append(s)
    return out[:2]


# 1. Supabase self-reported fascia names, by mobile.
supabase_fascia = {}
with open(SCRATCH / "exhibitors_fascia.json", encoding="utf-8") as f:
    for row in json.load(f):
        mobile = row.get("mobile")
        fj = row.get("fascia_names_json")
        if isinstance(fj, dict):
            names = fj.get("fascia_names") or []
        elif isinstance(fj, list):
            names = fj
        else:
            names = []
        cleaned = clean_names(names)
        if cleaned:
            supabase_fascia[mobile] = cleaned

# 2. Admin master report, by mobile.
admin_fascia = {}
wb = openpyxl.load_workbook(
    ROOT / "STE_2026_Admin_Master_Report_For_facia-names.xlsx", data_only=True
)
ws = wb["Exhibitor Master"]
header = [c.value for c in ws[1]]
col = {name: i for i, name in enumerate(header)}
for row in ws.iter_rows(min_row=2, values_only=True):
    mobile = row[col["Mobile (ID)"]]
    if not mobile:
        continue
    mobile = str(mobile).strip()
    raw = row[col["Fascia Names"]]
    if not raw:
        continue
    parts = [p.strip() for p in str(raw).split("|")]
    cleaned = clean_names(parts)
    if cleaned:
        admin_fascia[mobile] = cleaned

def norm_brand(name):
    return re.sub(r"[\s_./,()&-]+", "", str(name or "").lower())


# 3. master_exhibitor_stalls.csv: unitId + brand + every mobile that answers
#    for the firm (primary + aliases). Also collect every brand that already
#    has a real live stall somewhere, so the static-plan fallback below never
#    repeats a firm that has genuinely moved.
fascia_map = {}
unresolved = []
seated_brands = set()
unit_by_mobile = {}
with open(ROOT / "master_exhibitor_stalls.csv", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        unit_id = row["lottery_drawn_stall_number"] or row["portal_stall_number"]
        if not unit_id:
            continue  # not opened the Lucky Box yet - nothing live to letter

        seated_brands.add(norm_brand(row["brand"]))

        numbers = [row["primary_mobile"]] + [
            a for a in row["aliases"].split(";") if a
        ]
        for n in numbers:
            unit_by_mobile[n] = unit_id

        names = None
        source = None
        for n in numbers:
            if n in supabase_fascia:
                names = supabase_fascia[n]
                source = "supabase"
                break
        if not names:
            for n in numbers:
                if n in admin_fascia:
                    names = admin_fascia[n]
                    source = "admin_report"
                    break
        if not names:
            names = [row["brand"]]
            source = "brand_fallback"
            unresolved.append(row["brand"])

        fascia_map[unit_id] = {"fascia_names": names, "source": source}

# 4. Static plan (src/data/stallAllotment2026.ts): fills in a bay with no
#    live draw at all, as long as the firm it names has not already drawn a
#    different real stall (in which case this bay is genuinely empty now).
#
#    A `held: true` row is different in kind from the rest of the plan: it is
#    not a guess at where a firm might land, it is the organisers overriding
#    wherever the firm's own live draw already put them - so it always wins
#    over "this firm already drew a different real stall elsewhere". That
#    guard is only for the unheld rows, which are a simulated seeded draw
#    with no claim once a real one exists.
#
#    A held firm's old live bay is cleared with it - the organisers moved
#    them off it by hand, so it should stop showing their name. This is
#    matched by mobile (unit_by_mobile, built above from the same CSV) rather
#    than by brand string, because a firm's live fascia name is often not an
#    exact match for the plan's brand spelling - "Charchita Designer Sarees"
#    on the CSV against "Charchita Designer" on the plan, for one - and a
#    fuzzy match on brand text would either miss the stale entry or, worse,
#    clear the wrong one. A held bay some other firm has since drawn for real
#    is left showing that firm and printed as a conflict: the organisers'
#    plan and the live draw disagree, and that is a call for a human, not a
#    silent overwrite.
plan_text = (ROOT / "src" / "data" / "stallAllotment2026.ts").read_text(encoding="utf-8")
plan_pat = re.compile(
    r'unitId:\s*"([^"]+)",\s*stallNumber:\s*(\d+),\s*brand:\s*"([^"]*)"'
    r'[^{}]*?mobile:\s*"([^"]*)"'
    r'[^{}]*?held:\s*(true|false)'
)
plan_rows = [
    (unit_id, brand.strip(), mobile, held == "true")
    for unit_id, _stall_no, brand, mobile, held in plan_pat.findall(plan_text)
]

# Held rows are cleared in a pass of their own, before any placement, so a
# chain of moves (A's old bay is B's new one, B's old bay is C's new one)
# never depends on which order the bays happen to fall in the file. Without
# this, processing 14 (Charchita's new bay) before 30 (SANKALP's, whose old
# bay 14 has not been cleared yet) would see SANKALP's stale entry still
# sitting on 14 and wrongly call it a conflict.
for unit_id, brand, mobile, held in plan_rows:
    if not held or not brand or not mobile:
        continue
    old_unit = unit_by_mobile.get(mobile)
    if old_unit and old_unit != unit_id and old_unit in fascia_map:
        del fascia_map[old_unit]
        print(f"  held move: {brand} {old_unit} -> {unit_id} "
              f"(cleared stale live entry on {old_unit})")

plan_fallback_used = []
held_conflicts = []
for unit_id, brand, _mobile, held in plan_rows:
    if not brand:
        continue  # plan itself has no one down for this bay
    if unit_id in fascia_map:
        live_brand = fascia_map[unit_id]["fascia_names"][0]
        if held and norm_brand(live_brand) != norm_brand(brand):
            held_conflicts.append((unit_id, brand, live_brand))
        continue  # already has a live draw
    if not held and norm_brand(brand) in seated_brands:
        continue  # this firm already drew a different real stall elsewhere
    fascia_map[unit_id] = {
        "fascia_names": [brand],
        "source": "held_override" if held else "static_plan_fallback",
    }
    plan_fallback_used.append((unit_id, brand))

out_path = ROOT / "public" / "assets" / "fascia_map.json"
out_path.write_text(json.dumps(fascia_map, indent=2, ensure_ascii=False), encoding="utf-8")

print(f"wrote {len(fascia_map)} units to {out_path}")
print(f"  from supabase:        {sum(1 for v in fascia_map.values() if v['source'] == 'supabase')}")
print(f"  from admin report:    {sum(1 for v in fascia_map.values() if v['source'] == 'admin_report')}")
print(f"  brand-name fallback:  {len(unresolved)}")
print(f"  static-plan fallback: {len(plan_fallback_used)}")
if plan_fallback_used:
    for uid, brand in plan_fallback_used:
        print(f"    {uid}: {brand}")
if held_conflicts:
    print(f"  HELD-BAY CONFLICTS (need a human decision): {len(held_conflicts)}")
    for uid, planned, live in held_conflicts:
        print(f"    stall {uid}: plan wants {planned!r}, but {live!r} already holds it live")
